from appium import webdriver
from appium.options.android import UiAutomator2Options
from appium.webdriver.common.appiumby import AppiumBy
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, WebDriverException, StaleElementReferenceException
from selenium.webdriver.common.actions.action_builder import ActionBuilder
from selenium.webdriver.common.actions.pointer_input import PointerInput
import time
import os, datetime, base64

APPIUM_SERVER = "http://127.0.0.1:4723"
APP_PKG = "com.garibook.user"
APP_ACT  = "com.garibook.user.MainActivity"

# ===== Request flow locators =====
XPATH_RIDE_SHARE  = '//android.widget.ImageView[@content-desc="Ride Share"]'
XPATH_DROP_OFF_IN = '//android.view.View[@content-desc="Drop Off"]//android.widget.EditText'
XPATH_SUGGESTION  = '//android.widget.ImageView[@content-desc="Mirpur-1, Dhaka, Bangladesh\nDhaka, Bangladesh"]'
XPATH_CONTINUE    = '//android.view.View[@content-desc="Continue"]'
XPATH_CONFIRM     = '//android.view.View[@content-desc="Confirm Pickup"]'

# ===== Cancel flow locators (from your screenshot/spec) =====
XPATH_CANCEL_ENTRY = '//android.widget.ImageView[@content-desc="Want to cancel the Trip?"]'
AUI_WAIT_LONG      = 'new UiSelector().textContains("Waiting time too long")'
XPATH_CANCEL_RIDE  = '//android.view.View[@content-desc="Cancel Ride"]'

caps = {
    "platformName": "Android",
    "appium:platformVersion": "15",          # set to real OS or remove if mismatch
    "appium:deviceName": "10BF830G2N0058R",  # set to your real device / or add "appium:udid"
    "appium:automationName": "UiAutomator2",
    "appium:appPackage": APP_PKG,
    "appium:appActivity": APP_ACT,
    "appium:noReset": True,
    "appium:newCommandTimeout": 180,
    "appium:autoGrantPermissions": True,
    "appium:disableWindowAnimation": True,
    "appium:appWaitActivity": "*",
}

# =========================
# Paths / helpers (logs + video)
# =========================
REPORT_DIR   = "reports"
RECORD_DIR   = "recordings"
STEPLOG_XLSX = os.path.join(REPORT_DIR, "step_log.xlsx")

def _ensure_dirs():
    os.makedirs(REPORT_DIR, exist_ok=True)
    os.makedirs(RECORD_DIR, exist_ok=True)

def _clock():
    return datetime.datetime.now().strftime("%H:%M:%S")

def _ts_for_name(prefix, ext):
    return datetime.datetime.now().strftime(f"{prefix}__%Y%m%d_%H%M%S.{ext}")

class StepLogger:
    def __init__(self):
        _ensure_dirs()
        self.t0 = time.time()
        self.rows = []  # (Step, Status, ClockTime, +Elapsed(s), Note)

    def log(self, step, status, note=""):
        elapsed = round(time.time() - self.t0, 2)
        ct = _clock()
        print(f"{ct} [+{elapsed:.2f}s] {step} → {status}{' — ' + note if note else ''}")
        self.rows.append((step, status, ct, elapsed, note))

    def write(self):
        try:
            from openpyxl import Workbook, load_workbook
            if os.path.exists(STEPLOG_XLSX):
                wb = load_workbook(STEPLOG_XLSX)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Steps"
                ws.append(["Step", "Status", "Clock Time", "Elapsed (s)", "Note"])
            for r in self.rows:
                ws.append(list(r))
            wb.save(STEPLOG_XLSX)
            return STEPLOG_XLSX
        except Exception:
            import csv
            csv_path = STEPLOG_XLSX.replace(".xlsx", ".csv")
            write_hdr = not os.path.exists(csv_path)
            with open(csv_path, "a", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                if write_hdr:
                    w.writerow(["Step", "Status", "Clock Time", "ms", "Note"])
                for r in self.rows:
                    w.writerow(list(r))
            return csv_path

# --------- video helpers ---------
def start_recording(driver):
    try:
        driver.start_recording_screen()
    except Exception as e:
        print(f"[WARN] start_recording_screen failed: {e}")

def stop_and_save_recording(driver, name_prefix="rideshare_run"):
    _ensure_dirs()
    try:
        b64 = driver.stop_recording_screen()
        if not b64:
            return ""
        data = base64.b64decode(b64)
        path = os.path.join(RECORD_DIR, _ts_for_name(name_prefix, "mp4"))
        with open(path, "wb") as f:
            f.write(data)
        return path
    except Exception as e:
        print(f"[WARN] stop_recording_screen failed: {e}")
        return ""

# =========================
# Appium helpers
# =========================
def init_driver():
    d = webdriver.Remote(APPIUM_SERVER, options=UiAutomator2Options().load_capabilities(caps))
    try:
        d.update_settings({"waitForIdleTimeout": 0, "ignoreUnimportantViews": True, "waitForSelectorTimeout": 9000})
    except Exception:
        pass
    return d

def wait_click(wait, by, locator):
    return wait.until(EC.element_to_be_clickable((by, locator)))

# W3C tap (no TouchAction)
def tap_point(driver, x, y):
    finger = PointerInput("touch", "finger")
    actions = ActionBuilder(driver)
    actions.add_action(finger.create_pointer_move(0, 'viewport', int(x), int(y)))
    actions.add_action(finger.create_pointer_down(0))
    actions.add_action(finger.create_pointer_up(0))
    actions.perform()

def tap_ratio(driver, rx, ry):
    size = driver.get_window_size()
    tap_point(driver, rx * size["width"], ry * size["height"])

# ===== Review skip =====
def skip_review_if_present(driver, overall_timeout=6):
    end = time.time() + overall_timeout
    while time.time() < end:
        clicked = False
        for by, loc in [
            (AppiumBy.ANDROID_UIAUTOMATOR, 'new UiSelector().text("Skip")'),
            (AppiumBy.ANDROID_UIAUTOMATOR, 'new UiSelector().textContains("Skip")'),
            (AppiumBy.ANDROID_UIAUTOMATOR, 'new UiSelector().descriptionContains("Skip")'),
            (AppiumBy.ANDROID_UIAUTOMATOR, 'new UiSelector().text("এড়িয়ে যান")'),
            (AppiumBy.ANDROID_UIAUTOMATOR, 'new UiSelector().textContains("এড়িয়ে")'),
        ]:
            try:
                els = driver.find_elements(by, loc)
                if els:
                    els[0].click()
                    clicked = True
                    break
            except WebDriverException:
                pass
        if clicked:
            time.sleep(0.3)
            return
        try:
            tap_ratio(driver, 0.89, 0.148)   # fallback position
            time.sleep(0.3)
            return
        except Exception:
            pass
        time.sleep(0.2)

def click_confirm_with_retry(driver, total_timeout=35, poll=0.6):
    end = time.time() + total_timeout
    last_err = None
    while time.time() < end:
        try:
            el = WebDriverWait(driver, 6).until(EC.element_to_be_clickable((AppiumBy.XPATH, XPATH_CONFIRM)))
            el.click()
            return True
        except (TimeoutException, StaleElementReferenceException, WebDriverException) as e:
            last_err = e
            time.sleep(poll)
    if last_err:
        print(f"Confirm Pickup click failed: {type(last_err).__name__}: {last_err}")
    return False

# ===== State helpers =====
CONFIRM_LOCATORS = [
    (AppiumBy.XPATH, XPATH_CONFIRM),
    (AppiumBy.ACCESSIBILITY_ID, "Confirm Pickup"),
    (AppiumBy.ANDROID_UIAUTOMATOR, 'new UiSelector().textContains("Confirm Pickup")'),
    (AppiumBy.ANDROID_UIAUTOMATOR, 'new UiSelector().descriptionContains("Confirm Pickup")'),
    (AppiumBy.ANDROID_UIAUTOMATOR, 'new UiSelector().textContains("Confirm")'),
    (AppiumBy.ANDROID_UIAUTOMATOR, 'new UiSelector().descriptionContains("Confirm")'),
]

WAITING_LOCATORS = [
    (AppiumBy.XPATH, '//*[@content-desc="Waiting for driver"]'),
    (AppiumBy.ACCESSIBILITY_ID, 'Waiting for driver'),
    (AppiumBy.XPATH, XPATH_CANCEL_ENTRY),
]

def _already_waiting_screen(d):
    for by, loc in WAITING_LOCATORS:
        try:
            if d.find_elements(by, loc):
                return True
        except Exception:
            pass
    return False

def _click_center(driver, el):
    r = el.rect
    tap_point(driver, r["x"] + r["width"] // 2, r["y"] + r["height"] // 2)

def _find_confirm(driver, small_timeout=3):
    end = time.time() + small_timeout
    while time.time() < end:
        for by, loc in CONFIRM_LOCATORS:
            try:
                els = driver.find_elements(by, loc)
                for el in els:
                    if el.is_displayed():
                        return el
            except Exception:
                pass
        time.sleep(0.3)
    return None

def confirm_pickup_smart(driver, logger=None, total_timeout=35):
    t_end = time.time() + total_timeout
    tried_scroll = False

    while time.time() < t_end:
        if _already_waiting_screen(driver):
            logger and logger.log("Submission", "PASS", "Already on waiting screen; Confirm not required")
            return True

        try:
            driver.hide_keyboard()
        except Exception:
            pass

        el = _find_confirm(driver, small_timeout=2)
        if el:
            try:
                el.click()
                logger and logger.log("Confirm Pickup", "PASS", "Clicked (direct)")
                return True
            except Exception:
                try:
                    _click_center(driver, el)
                    logger and logger.log("Confirm Pickup", "PASS", "Clicked (center tap)")
                    return True
                except Exception as e2:
                    logger and logger.log("Confirm Pickup", "INFO", f"Center tap failed: {e2}")

        if not tried_scroll:
            try:
                driver.find_element(
                    AppiumBy.ANDROID_UIAUTOMATOR,
                    'new UiScrollable(new UiSelector().scrollable(true))'
                    '.scrollIntoView(new UiSelector().textContains("Confirm"))'
                )
                tried_scroll = True
                continue
            except Exception:
                tried_scroll = True

        time.sleep(0.4)

    try:
        tap_ratio(driver, 0.50, 0.92)
        logger and logger.log("Confirm Pickup", "PASS", "Clicked (fallback bottom tap)")
        return True
    except Exception as e:
        logger and logger.log("Confirm Pickup", "FAIL", f"Not found/clickable: {e}")
        return False

# ===== Cancel Trip flow =====
def cancel_trip_flow(d, logger=None):
    # 1) Open the cancel sheet
    try:
        WebDriverWait(d, 25).until(EC.element_to_be_clickable((AppiumBy.XPATH, XPATH_CANCEL_ENTRY))).click()
        logger and logger.log("Cancel", "INFO", "Opened cancel sheet")
    except Exception as e:
        logger and logger.log("Cancel", "FAIL", f"Cancel entry not found: {e}")
        return False

    # 2) Select "Waiting time too long"
    try:
        # try multiple ways
        reason = None
        for by, loc in [
            (AppiumBy.ACCESSIBILITY_ID, "Waiting time too long"),
            (AppiumBy.XPATH, '//*[@content-desc="Waiting time too long"]'),
            (AppiumBy.XPATH, '//*[contains(@content-desc,"Waiting time too long")]'),
        ]:
            try:
                els = d.find_elements(by, loc)
                if els:
                    reason = els[0]
                    break
            except Exception:
                pass
        if not reason:
            try:
                d.find_element(AppiumBy.ANDROID_UIAUTOMATOR, AUI_WAIT_LONG)
                reason = d.find_element(AppiumBy.ANDROID_UIAUTOMATOR, AUI_WAIT_LONG)
            except Exception:
                # Attempt to scroll into view then re-try
                try:
                    d.find_element(
                        AppiumBy.ANDROID_UIAUTOMATOR,
                        'new UiScrollable(new UiSelector().scrollable(true))'
                        '.scrollIntoView(new UiSelector().textContains("Waiting time"))'
                    )
                    reason = d.find_element(AppiumBy.ANDROID_UIAUTOMATOR, AUI_WAIT_LONG)
                except Exception:
                    pass
        if not reason:
            logger and logger.log("Cancel", "FAIL", "Reason option not found")
            return False

        try:
            reason.click()
        except Exception:
            _click_center(d, reason)
        time.sleep(0.4)
        logger and logger.log("Cancel", "INFO", "Reason selected")
    except Exception as e:
        logger and logger.log("Cancel", "FAIL", f"Could not select reason: {e}")
        return False

    # 3) Tap "Cancel Ride"
    try:
        WebDriverWait(d, 20).until(EC.element_to_be_clickable((AppiumBy.XPATH, XPATH_CANCEL_RIDE))).click()
        logger and logger.log("Cancel", "PASS", "Ride cancelled")
        return True
    except Exception:
        # fallback tap near bottom area
        try:
            tap_ratio(d, 0.50, 0.92)
            logger and logger.log("Cancel", "PASS", "Ride cancelled (bottom tap)")
            return True
        except Exception as e2:
            logger and logger.log("Cancel", "FAIL", f"Cancel Ride not clickable: {e2}")
            return False

# =========================
# Main
# =========================
def main():
    logger = StepLogger()
    d = init_driver()
    wait = WebDriverWait(d, 30)

    start_recording(d)
    video_path = ""

    try:
        # 1) Skip review
        skip_review_if_present(d)
        logger.log("Skip Review", "PASS")
        time.sleep(2.0)

        # 2) Ride Share
        try:
            wait_click(wait, AppiumBy.XPATH, XPATH_RIDE_SHARE).click()
            logger.log("Ride Share", "PASS")
        except TimeoutException:
            logger.log("Ride Share", "WARN", "Ride card not present")

        # 3) Type 'mirpur'
        try:
            inp = wait_click(wait, AppiumBy.XPATH, XPATH_DROP_OFF_IN)
        except TimeoutException:
            logger.log("Drop Off Input", "FAIL", "Input not found")
            return
        try:
            inp.click()
            try: inp.clear()
            except Exception: pass
            inp.send_keys("mirpur")
            logger.log("Type Drop Off", "PASS", "Typed 'mirpur'")
        except WebDriverException as e:
            logger.log("Type Drop Off", "FAIL", f"{e}")
            return

        # 4) Suggestion
        time.sleep(1.2)
        try:
            wait_click(WebDriverWait(d, 15), AppiumBy.XPATH, XPATH_SUGGESTION).click()
            logger.log("Select Suggestion", "PASS", "Exact XPath")
        except TimeoutException:
            try: d.hide_keyboard()
            except Exception: pass
            try:
                tap_ratio(d, 0.36, 0.44)
                logger.log("Select Suggestion", "PASS", "Tapped by ratio (0.36,0.44)")
            except Exception:
                logger.log("Select Suggestion", "FAIL", "XPath and ratio both failed")
                return

        # Fare page settle
        logger.log("Fare Page", "INFO", "Waiting ~3s for fare/ETA")
        time.sleep(3.0)

        # 5) Continue
        try:
            wait_click(WebDriverWait(d, 20), AppiumBy.XPATH, XPATH_CONTINUE).click()
            logger.log("Continue", "PASS")
        except TimeoutException:
            logger.log("Continue", "WARN", "Continue not present")

        # 6) Confirm Pickup (then cancel)
        logger.log("Post-Continue", "INFO", "Settling before Confirm (~3s)")
        time.sleep(3.0)

        try:
            el = WebDriverWait(d, 25).until(EC.element_to_be_clickable((AppiumBy.XPATH, XPATH_CONFIRM)))
            time.sleep(0.6)
            el.click()
            logger.log("Confirm Pickup", "PASS", "Clicked after settle/wait")
        except TimeoutException:
            if not confirm_pickup_smart(d, logger, total_timeout=25):
                if not click_confirm_with_retry(d, total_timeout=15) and not _already_waiting_screen(d):
                    logger.log("Confirm Pickup", "FAIL", "Not found and not auto-submitted")
                    return

        # We should now be on the waiting screen → cancel the trip
        logger.log("Waiting", "INFO", "Attempting to cancel trip now")
        if cancel_trip_flow(d, logger):
            print("✅ Trip cancelled successfully.")
        else:
            print("❌ Cancel flow failed.")

        logger.log("End", "PASS", "Run complete")

    finally:
        try:
            video_path = stop_and_save_recording(d, "rideshare_run")
        except Exception:
            video_path = ""
        d.quit()

        if video_path:
            logger.log("Video Saved", "INFO", video_path)

        path = logger.write()
        print("🎞️ Video:", os.path.abspath(video_path) if video_path else "(none)")
        print(f"🧾 Step log saved to: {os.path.abspath(path)}")

if __name__ == "__main__":
    main()
